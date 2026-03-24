"""
Build 语法对应关系表.xlsx
Primary: ACTFL grammar → match HSK (File3) and PG (File2)
Appendix: unmatched HSK/PG items appended at bottom
"""
import pandas as pd
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE     = r"D:\2026 ACTFL 词表\词表202603\结项报告"
ACTFL_FILE = BASE + r"\ACTFL语法量表总表.xlsx"
PG_FILE    = BASE + r"\等级标准语法大纲.xlsx"
HSK_FILE   = BASE + r"\HSK_Grammar_MCP_Resource_Cleaned.csv"
OUT_FILE   = BASE + r"\语法对应关系表.xlsx"

# ── Load ───────────────────────────────────────────────────────────────────────
actfl = pd.read_excel(ACTFL_FILE)
pg    = pd.read_excel(PG_FILE)
hsk   = pd.read_csv(HSK_FILE, encoding="utf-8-sig")
actfl.columns = actfl.columns.str.strip()
pg.columns    = pg.columns.str.strip()
hsk.columns   = hsk.columns.str.strip()

# ── Level ordering ─────────────────────────────────────────────────────────────
LEVEL_NUM = {"一级":1,"二级":2,"三级":3,"四级":4,"五级":5,"六级":6,
             "七级":7,"七一九级":8,"七—九级":8,"七－九级":8}
def lvl_key(s): return LEVEL_NUM.get(str(s).strip(), 99)

# ── ACTFL→HSK/PG name aliases (where naming conventions differ) ───────────────
# Maps ACTFL extracted key → list of alternative search terms to try
ALIASES = {
    "被字句":    ["被动句"],
    "趋向动词":  ["趋向补语"],
    "方位名词短语": ["方位名词","方位短语"],
    "处所词":    ["处所"],
    "形容词谓语句": ["形容词谓语","形容词性谓语"],
    "名词谓语句":   ["名词性谓语","名词谓语"],
    "动词谓语句":   ["动词谓语"],
    "主谓谓语句":   ["主谓谓语"],
    "感叹句":    ["感叹"],
    "特指问":    ["特指疑问句","疑问代词"],
    "是非问":    ["是非疑问句"],
    "正反问":    ["正反疑问句"],
    "选择问":    ["选择疑问句"],
    "语气助词":  ["语气词"],
    "结构助词的": ["结构助词"],
    "离合动词":  ["离合词"],
    "程度副词":  ["程度"],
    "兼语句":    ["兼语"],
    "连动句":    ["连动"],
    "存现句":    ["存现"],
    "双宾语":    ["双宾"],
    "补语":      ["补语"],
}

# ── Key-term extraction from ACTFL grammar item text ──────────────────────────
def extract_key(text: str) -> str:
    t = str(text).strip()
    t = re.sub(r"^\d*\s*【[^】]+】\s*", "", t)    # remove 【N】 or 425【001】
    t = t.split("：")[0].split(":")[0].strip()       # take part before colon
    t = re.sub(r"[（(][^）)]+[)）]", "", t).strip()  # remove （参见...）
    for q in ('"', '"', '"', '"', '"', "'"):
        t = t.replace(q, "")
    t = re.sub(r"\.{2,}", "", t).strip()             # remove ...
    t = t.strip("—–-·· ")
    t = re.sub(r"\d+\s*$", "", t).strip()            # trailing number LAST
    return t.strip()

# ── Generic find: returns (earliest_level, all_matched_indices) ────────────────
def find_all(search_terms, df, exact_cols, fuzzy_cols, level_col="等级"):
    """Try each term in search_terms; return (earliest_level, set_of_indices)."""
    all_idx = set()
    best_lv = None

    for key in search_terms:
        if not key or len(key) < 2:
            continue
        esc = re.escape(key)
        found = False

        for col in exact_cols:
            if col not in df.columns: continue
            m = df[col].fillna("").str.strip() == key
            if m.any():
                hits = df[m]
                if best_lv is None or lvl_key(hits.iloc[0][level_col]) < lvl_key(best_lv or ""):
                    best_lv = str(hits.iloc[0][level_col])
                all_idx.update(hits.index.tolist())
                found = True

        if not found:
            for col in exact_cols:
                if col not in df.columns: continue
                m = df[col].fillna("").str.contains(esc, regex=True)
                if m.any():
                    hits = df[m]
                    if best_lv is None or lvl_key(hits.iloc[0][level_col]) < lvl_key(best_lv or ""):
                        best_lv = str(hits.iloc[0][level_col])
                    all_idx.update(hits.index.tolist())
                    found = True
                    break

        if not found:
            for col in fuzzy_cols:
                if col not in df.columns: continue
                m = df[col].fillna("").str.contains(esc, regex=True)
                if m.any():
                    hits = df[m]
                    if best_lv is None or lvl_key(hits.iloc[0][level_col]) < lvl_key(best_lv or ""):
                        best_lv = str(hits.iloc[0][level_col])
                    all_idx.update(hits.index.tolist())
                    break

    return best_lv, all_idx


def search_terms_for(key):
    """Build list of terms to try: key itself + any aliases."""
    terms = [key]
    if key in ALIASES:
        terms += ALIASES[key]
    return terms


# ── Pass 1: ACTFL items ────────────────────────────────────────────────────────
pg_matched  = set()
hsk_matched = set()
rows_main   = []

for _, ar in actfl.iterrows():
    item_full  = str(ar.get("语法项", ""))
    key        = extract_key(item_full)
    gram_class = str(ar.get("语法类", "")).strip()
    terms      = search_terms_for(key)

    pg_lv,  pi_set = find_all(terms, pg,  ["细目","类别名称"], ["语法内容"])
    hsk_lv, hi_set = find_all(terms, hsk, ["细目","类别名称"], ["语法内容"])
    pg_matched.update(pi_set)
    hsk_matched.update(hi_set)

    parts = ["ACTFL"]
    if hsk_lv: parts.append("HSK")
    if pg_lv:  parts.append("PG")

    rows_main.append({
        "语法项":    key,
        "语法类":    gram_class,
        "ACTFL等级": str(ar.get("等级", "")),
        "HSK等级":   hsk_lv or "",
        "PG等级":    pg_lv  or "",
        "来源":      "+".join(parts),
        "语法描述":  item_full,
        "例句":      str(ar.get("例句", "")) if pd.notna(ar.get("例句")) else "",
    })

# ── Pass 2: unmatched PG → check vs unmatched HSK ─────────────────────────────
unmatched_pg  = pg[~pg.index.isin(pg_matched)].copy()
unmatched_hsk = hsk[~hsk.index.isin(hsk_matched)].copy()

hsk2_matched = set()
rows_app     = []

for idx, pr in unmatched_pg.iterrows():
    key = str(pr.get("细目", "") or "").strip()
    if not key or key == "nan":
        key = str(pr.get("类别名称","") or "").strip()
    if key == "nan": key = ""

    terms = search_terms_for(key)
    hsk_lv2, hi2_set = find_all(terms, unmatched_hsk,
                                 ["细目","类别名称"], ["语法内容"])
    hsk2_matched.update(hi2_set)

    source  = "HSK+PG共有" if hsk_lv2 else "PG独有"
    cat     = f"{str(pr.get('类别',''))}{str(pr.get('类别名称',''))}".strip()
    content = str(pr.get("细目","") or pr.get("语法内容","") or "").strip()
    if content == "nan": content = ""
    example = str(pr.get("例句","") or "").strip()
    if example == "nan": example = ""

    rows_app.append({
        "语法项":    key or content[:30],
        "语法类":    cat,
        "ACTFL等级": "",
        "HSK等级":   hsk_lv2 or "",
        "PG等级":    str(pr.get("等级","")),
        "来源":      source,
        "语法描述":  content,
        "例句":      example,
        "_sort": ({"HSK+PG共有":0,"PG独有":1,"HSK独有":2}.get(source,9),
                  lvl_key(pr.get("等级",""))),
    })

# HSK-only remaining
for idx, hr in unmatched_hsk[~unmatched_hsk.index.isin(hsk2_matched)].iterrows():
    key = str(hr.get("细目","") or "").strip()
    if not key or key == "nan":
        key = str(hr.get("类别名称","") or "").strip()
    if key == "nan": key = ""
    cat     = f"{str(hr.get('类别',''))}{str(hr.get('类别名称',''))}".strip()
    content = str(hr.get("语法内容","") or "").strip()
    if content == "nan": content = ""

    rows_app.append({
        "语法项":    key or content[:30],
        "语法类":    cat,
        "ACTFL等级": "",
        "HSK等级":   str(hr.get("等级","")),
        "PG等级":    "",
        "来源":      "HSK独有",
        "语法描述":  content,
        "例句":      "",
        "_sort": (2, lvl_key(hr.get("等级",""))),
    })

rows_app.sort(key=lambda r: r["_sort"])
for r in rows_app: r.pop("_sort", None)

all_rows = rows_main + rows_app

# ── Stats ──────────────────────────────────────────────────────────────────────
print(f"Main (ACTFL): {len(rows_main)}")
matched_any = sum(1 for r in rows_main if "+" in r["来源"])
match_hsk   = sum(1 for r in rows_main if "HSK" in r["来源"])
match_pg    = sum(1 for r in rows_main if "PG"  in r["来源"])
print(f"  匹配HSK: {match_hsk}  匹配PG: {match_pg}  任意匹配: {matched_any}")
print(f"Appendix: {len(rows_app)}")
print(f"  HSK+PG共有: {sum(1 for r in rows_app if r['来源']=='HSK+PG共有')}")
print(f"  PG独有:     {sum(1 for r in rows_app if r['来源']=='PG独有')}")
print(f"  HSK独有:    {sum(1 for r in rows_app if r['来源']=='HSK独有')}")
print(f"Total rows: {len(all_rows)}")

print("\n=== Sample ACTFL matches ===")
for r in rows_main[:8]:
    print(f"  {r['语法项'][:16]:16} | {r['ACTFL等级'][:16]:16} | HSK:{r['HSK等级']:6} | PG:{r['PG等级']:6} | {r['来源']}")

print("\n=== Sample appendix ===")
for r in rows_app[:8]:
    print(f"  {r['语法项'][:16]:16} | {'---':16} | HSK:{r['HSK等级']:6} | PG:{r['PG等级']:6} | {r['来源']}")

# ── Write Excel ────────────────────────────────────────────────────────────────
COLS    = ["语法项","语法类","ACTFL等级","HSK等级","PG等级","来源","语法描述","例句"]
HEADERS = ["语法项","语法类","ACTFL等级","HSK等级","PG等级","来源",
           "语法描述（ACTFL原文/HSK/PG内容）","例句"]
WIDTHS  = [20, 26, 18, 10, 10, 15, 48, 52]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "语法对应关系"

thin  = Side(style="thin",  color="CCCCCC")
med   = Side(style="medium", color="4472C4")
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
b_sep = Border(left=thin, right=thin, top=med,  bottom=thin)

hdr_font  = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
hdr_fill  = PatternFill("solid", fgColor="2D5FA0")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="宋体", size=9)
cent_top  = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_top  = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# Per-source subtle row tint
SRC_FILL = {
    "ACTFL":        PatternFill("solid", fgColor="FFFFFF"),
    "ACTFL+HSK":    PatternFill("solid", fgColor="EDF3FB"),
    "ACTFL+PG":     PatternFill("solid", fgColor="EDFBEE"),
    "ACTFL+HSK+PG": PatternFill("solid", fgColor="E6F5E6"),
    "HSK+PG共有":   PatternFill("solid", fgColor="FFFAE6"),
    "PG独有":       PatternFill("solid", fgColor="FFF3E6"),
    "HSK独有":      PatternFill("solid", fgColor="F3EEFF"),
}

# Header
for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = hdr_align; c.border = b_all
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

sep_row = len(rows_main) + 2  # first appendix row

for ri, row_data in enumerate(all_rows, 2):
    src  = row_data["来源"]
    fill = SRC_FILL.get(src, PatternFill("solid", fgColor="FFFFFF"))
    use_sep = (ri == sep_row)

    for ci, col in enumerate(COLS, 1):
        val = row_data.get(col, "")
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = body_font; c.fill = fill
        c.border = b_sep if use_sep else b_all
        c.alignment = cent_top if ci <= 6 else left_top

    ex   = str(row_data.get("例句", ""))
    desc = str(row_data.get("语法描述",""))
    lines = max(1, max(len(ex), len(desc)) // 25 + ex.count("\n") + 1)
    ws.row_dimensions[ri].height = max(15, min(lines * 14, 110))

# ── Legend sheet ───────────────────────────────────────────────────────────────
leg = wb.create_sheet("说明")
leg_data = [
    ["来源标记", "含义", "行背景色"],
    ["ACTFL",         "仅出现在ACTFL语法量表中",            "白色"],
    ["ACTFL+HSK",     "ACTFL与HSK均有收录",                "浅蓝"],
    ["ACTFL+PG",      "ACTFL与等级标准（PG）均有收录",      "浅绿"],
    ["ACTFL+HSK+PG",  "三个体系均有收录",                   "深浅绿"],
    ["HSK+PG共有",    "HSK与PG共有，ACTFL未收录（附录）",   "浅黄"],
    ["PG独有",        "仅等级标准收录（附录）",              "浅橙"],
    ["HSK独有",       "仅HSK收录（附录）",                  "浅紫"],
    [],
    ["等级对照参考（仅供参考，非官方对应）"],
    ["ACTFL等级", "大致对应HSK/PG等级"],
    ["Novice Low/Mid",       "HSK 1级"],
    ["Novice High",          "HSK 2-3级"],
    ["Intermediate Low",     "HSK 3-4级"],
    ["Intermediate Mid",     "HSK 4-5级"],
    ["Intermediate High",    "HSK 5-6级"],
    ["Advanced+",            "HSK 七—九级"],
    [],
    ["数据来源"],
    ["ACTFL",      "ACTFL语法量表总表.xlsx（本项目整理）"],
    ["HSK",        "HSK_Grammar_MCP_Resource_Cleaned.csv"],
    ["PG（等级标准）","等级标准语法大纲.xlsx（国际中文教育中文水平等级标准）"],
    [],
    ["匹配说明"],
    ["• 以ACTFL语法项为主体，自动搜索HSK/PG中的对应项"],
    ["• 匹配基于关键词提取，包含别名映射（如\"被字句\"对应\"被动句\"）"],
    ["• HSK/PG中所有子条目均统计覆盖，确保同名语法点不重复出现于附录"],
    ["• 附录部分按 HSK+PG共有 → PG独有 → HSK独有 排序，各组内按等级升序"],
]
for row in leg_data:
    leg.append(row)
leg.column_dimensions["A"].width = 20
leg.column_dimensions["B"].width = 40
leg.column_dimensions["C"].width = 12
for r in [1, 10, 19, 24]:
    if leg.cell(r, 1).value:
        leg.cell(r, 1).font = Font(bold=True, color="2D5FA0")

wb.save(OUT_FILE)
print(f"\nSaved → {OUT_FILE}")
