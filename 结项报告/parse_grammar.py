"""
Parse 语法量表总表.docx → Excel
Columns: 等级 | 语法类 | 编号 | 语法项 | 例句
"""
import re
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DOCX_XML = r"D:\2026 ACTFL 词表\词表202603\结项报告\unpacked_grammar\word\document.xml"
OUTPUT   = r"D:\2026 ACTFL 词表\词表202603\结项报告\语法量表总表_v3.xlsx"

NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

LEVEL_PATTERN = re.compile(
    r"(Novice|Intermediate|Advanced|Superior|Distinguished)"
    r"[\s\-–—]*(Low|Mid|High|Plus|\+)?",
    re.IGNORECASE
)

# Item patterns:
#   Standard:   【11】xxx  /  【1a】xxx
#   Advanced+:  425【 001】 xxx  /  425【·147】 xxx
ITEM_PATTERN = re.compile(r"^(\d+\s*)?【[\s·\d]+[a-zA-Z]?】")


# ── XML helpers ────────────────────────────────────────────────────────────────

def get_text(p):
    parts = []
    for r in p.findall(".//w:r", NS):
        for t in r.findall("w:t", NS):
            if t.text:
                parts.append(t.text)
    return "".join(parts).strip()


def get_ea_fonts(p):
    fonts = set()
    for r in p.findall(".//w:r", NS):
        rpr = r.find("w:rPr", NS)
        if rpr is not None:
            rf = rpr.find("w:rFonts", NS)
            if rf is not None:
                ea = rf.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}eastAsia")
                if ea:
                    fonts.add(ea)
    return fonts


def is_bold(p):
    ppr = p.find("w:pPr", NS)
    if ppr is not None:
        rpr = ppr.find("w:rPr", NS)
        if rpr is not None and rpr.find("w:b", NS) is not None:
            return True
    for r in p.findall("w:r", NS):
        rpr = r.find("w:rPr", NS)
        if rpr is not None and rpr.find("w:b", NS) is not None:
            return True
    return False


def get_firstline_indent(p):
    ppr = p.find("w:pPr", NS)
    if ppr is not None:
        ind = ppr.find("w:ind", NS)
        if ind is not None:
            fl = ind.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}firstLine")
            if fl:
                return int(fl)
    return 0


def is_outline_0(p):
    ppr = p.find("w:pPr", NS)
    if ppr is not None:
        ol = ppr.find("w:outlineLvl", NS)
        if ol is not None:
            return ol.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val") == "0"
    return False


def has_large_font(p):
    for r in p.findall(".//w:r", NS):
        rpr = r.find("w:rPr", NS)
        if rpr is not None:
            sz = rpr.find("w:sz", NS)
            if sz is not None:
                val = sz.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val")
                if val and int(val) >= 28:
                    return True
    return False


# ── Item text splitting ────────────────────────────────────────────────────────

def split_item_example(text):
    """
    If an item paragraph has example sentences appended to the item description,
    split them apart.
    Returns (item_desc, inline_example_or_empty).
    Example: '572【148】偏正关系：...让步句群 他们俩正是这样...'
          -> ('572【148】偏正关系：...让步句群', '他们俩正是这样...')
    """
    if not re.search(r"[。？！]", text):
        return text, ""

    # Find last half-width or full-width space before the first sentence-ending char
    first_end = re.search(r"[。？！]", text).start()
    pre = text[:first_end]

    # Look for a space (or full-width space) that separates item vocab from sentence
    for sep in (" ", "\u3000"):
        idx = pre.rfind(sep)
        if idx > 5:
            candidate = text[idx:].strip()
            # Make sure candidate starts a sentence (not a bracket or punctuation)
            if candidate and candidate[0] not in "【（(、，：；":
                return text[:idx].strip(), candidate
    return text, ""


# ── Number extraction ──────────────────────────────────────────────────────────

def extract_number(text):
    """
    Extract the display number from an item heading.
    '【11】动词谓语句'        → '11'
    '425【 001】 能愿动词：需' → '425'
    '【1a】...'               → '1a'
    """
    # Try outer number first (Advanced+)
    m = re.match(r"^(\d+)\s*【", text)
    if m:
        return m.group(1)
    # Inner number
    m = re.match(r"^【[\s·]*(\d+[a-zA-Z]?)[\s·]*】", text)
    if m:
        return m.group(1)
    return ""


# ── Paragraph classifier ───────────────────────────────────────────────────────

def classify(p):
    """Return ('kind', text). Kinds: level, category, subcategory, item, example, skip"""
    text = get_text(p)
    if not text:
        return "skip", text

    fonts  = get_ea_fonts(p)
    indent = get_firstline_indent(p)
    bold   = is_bold(p)

    # ── Level ──────────────────────────────────────────────────────────────────
    if LEVEL_PATTERN.search(text) and (
            is_outline_0(p) or has_large_font(p) or
            not any("\u4e00" <= c <= "\u9fff" for c in text)):
        return "level", text

    # ── Item ───────────────────────────────────────────────────────────────────
    if ITEM_PATTERN.match(text):
        return "item", text

    # ── Category / Sub-category ────────────────────────────────────────────────
    if bold:
        if re.match(r"^[一二三四五六七八九十百]+[、．.]", text):
            return "category", text
        # Sub-category with Chinese or ASCII parentheses: （一）/ (一) / (一）
        if re.match(r"^[（(][一二三四五六七八九十百]+[)）]", text):
            return "subcategory", text
        # Short bold label without 【】 or （）, e.g. "语体", "篇章"
        if "【" not in text and len(text) <= 20 and indent == 0:
            return "category", text

    # ── Example ────────────────────────────────────────────────────────────────
    # 楷体 font → always example
    if "楷体" in fonts:
        return "example", text

    # Any indented paragraph → example
    if indent >= 400:
        return "example", text

    # No CJK font info + not bold → Advanced+-style running example
    if not fonts and not bold:
        return "example", text

    return "skip", text


# ── Document parser ────────────────────────────────────────────────────────────

def parse_document():
    tree = ET.parse(DOCX_XML)
    body = tree.getroot().find("w:body", NS)

    rows = []
    current_level    = ""
    current_category = ""
    current_sub      = ""
    current_item     = ""
    current_number   = ""
    current_examples = []

    def flush():
        if current_item:
            grammar_class = current_category
            if current_sub:
                grammar_class = f"{current_category}  {current_sub}" if current_category else current_sub
            rows.append((current_level, grammar_class, current_number,
                         current_item, "\n".join(current_examples)))

    for p in body.findall("w:p", NS):
        kind, text = classify(p)

        if kind == "level":
            flush()
            current_level = text
            current_category = current_sub = current_item = current_number = ""
            current_examples = []

        elif kind == "category":
            flush()
            current_category = text
            current_sub = current_item = current_number = ""
            current_examples = []

        elif kind == "subcategory":
            flush()
            current_sub = text
            current_item = current_number = ""
            current_examples = []

        elif kind == "item":
            flush()
            item_desc, inline_ex = split_item_example(text)
            current_item     = item_desc
            current_number   = extract_number(item_desc)
            current_examples = [inline_ex] if inline_ex else []

        elif kind == "example":
            if current_item:
                current_examples.append(text)

    flush()
    return rows


# ── Excel writer ───────────────────────────────────────────────────────────────

def write_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "语法量表"

    headers    = ["等级", "语法类", "编号", "语法项", "例句"]
    col_widths = [18, 28, 8, 48, 55]

    # Header row
    header_font  = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill(fill_type="solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

    # Body rows — plain white, no colour bands
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_top   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    plain_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    for ri, (level, grammar_class, number, item, examples) in enumerate(rows, 2):
        values = [level, grammar_class, number, item, examples]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = plain_fill
            cell.border = border
            cell.font   = Font(name="宋体", size=10)
            cell.alignment = center_top if ci <= 3 else left_top

        n_lines = examples.count("\n") + 1 if examples else 1
        ws.row_dimensions[ri].height = max(18, min(n_lines * 17, 120))

    wb.save(OUTPUT)
    print(f"Saved {len(rows)} rows → {OUTPUT}")


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    rows = parse_document()
    print(f"Parsed {len(rows)} grammar items.\n")

    # Spot-checks
    print("=== First 3 rows ===")
    for r in rows[:3]:
        print(r)

    print("\n=== 被字句2 ===")
    for r in rows:
        if "被" in r[3] and "字句2" in r[3]:
            print(r)

    print("\n=== 571/572 ===")
    for r in rows:
        if r[2] in ("571", "572"):
            print(f"  item: {r[3][:60]}")
            print(f"  ex:   {r[4][:120]}")

    print(f"\n=== Advanced+ total: {sum(1 for r in rows if 'Advanced' in r[0])} ===")

    write_excel(rows)
